import re
import sys
import datetime
import requests
import time, os, tarfile
import openpyxl
import pyodbc


from .MailUtil import send_email
from .StringUtil import tripString, isEmptyOrNone
from dvadmin.system.views.xingqi.models.Material import Material, session, engine
from dvadmin.system.views.xingqi.models.MaterialPrice import MaterialPrice
from dvadmin.system.views.xingqi.models.MaterialPriceSummary import MaterialPriceSummary
from sqlalchemy import text

from dvadmin.system.util.sql_config import brand_dic, mode_dic

import logging
logger = logging.getLogger(__name__)

#处理报价文件
def handleMaterialPrice(filename, file_md5):
    #排查文件是否上传过，如果上传提示用户
    mps = session.query(MaterialPriceSummary).filter(MaterialPriceSummary.file_md5 == file_md5).all()
    if len(mps) > 0:
        logger.info("已报过价了,不要重复上传。")
        return {'code': -1, 'errmsg': '已报过价了,不要重复上传。'}
    file_name = filename
    refer_excel = openpyxl.load_workbook(file_name)
    # 获取第一个sheet表格 报错后回复给前端提醒
    sheet1 = refer_excel['比价表']
    material_list = []

    #处理物料 判断mode和brand是否正确？需要反馈给前端，如果不对则修改
    mode_error_info = []
    brand_error_info = []
    for row in range(6, sheet1.max_row + 1):

        mode = tripString((sheet1.cell(row=row, column=2)).value)
        brand = tripString((sheet1.cell(row=row, column=5)).value)
        dic = {
            "xingqi_number": (sheet1.cell(row=row, column=3)).value,
            "name": (sheet1.cell(row=row, column=4)).value,
            "number": tripString((sheet1.cell(row=row, column=6)).value),
            "mode": mode,
            "brand": brand,
            "supplier": (sheet1.cell(row=4, column=14)).value,
            "amount": (sheet1.cell(row=row, column=7)).value,
            "price": (sheet1.cell(row=row, column=14)).value,
        }
        material_list.append(dic)
        if isEmptyOrNone(mode) or isEmptyOrNone(brand):
            mode_error_info.append("第"+str(row) +  "类型或品牌是空的")
        else:
            if isEmptyOrNone(mode_dic.get(mode)):
                mode_error_info.append("第"+str(row) + "行数据类型是空的")
            if isEmptyOrNone(brand_dic.get(brand)):
                brand_error_info.append("第"+str(row) + '行数据品牌是空的')

    if len(mode_error_info) > 0 or len(brand_error_info) > 0:
        logger.info("error--------物料品牌或者类型不能是空")
        logger.info(mode_error_info)
        logger.info(brand_error_info)
        return {
        'code': 2593,
        'errmsg': '物料品牌或者类型不能是空',
    }

    for material in material_list:
        r1 = session.query(Material).filter(Material.material_number == material["number"]).all()
        if len(r1) > 0:
            #说明存在了
            logger.info("这个料存在"+material["number"])
        else:
            m = Material(material_number=material.get("number"),
                         material_name=material.get("name"),
                                material_brand=material.get("brand"),
                         material_mode=material.get("mode"),
                         xingqi_number=material.get("xingqi_number"),
                         )
            session.add(m)
            print(m.material_id)
    session.commit()
    #先增加物料信息
    return {
        'code': 0,
        'errmsg': 'ok',
        'material_list': material_list,
        'file_md5': file_md5,
        'filename': filename,
    }

#存储报价的介绍信息
def saveMaterialPriceSummary(priceSummaryInfo):
    mps = MaterialPriceSummary(supplier = priceSummaryInfo.get("supplier"),
                               creator = 'leo',
                               info = priceSummaryInfo.get("info"),
                               filename = priceSummaryInfo.get("filename"),
                               file_md5 = priceSummaryInfo.get("file_md5"))
    session.add(mps)
    session.commit()
    sql = text('SELECT LAST_INSERT_ID();')
    result_id = engine.execute(sql).all()[0][0]
    saveMaterialPriceList(priceSummaryInfo['list'], result_id)

#根据brand和mode下载系统中的价格信息
def downloadMaterialPriceWithBrandAndMode(brandModeInfo):
    logger.info(brandModeInfo.get("brand"))
    sql = text("""
    
    select m.material_number, m.material_name, m.material_mode, m.material_brand, mp.price, mp.amount, mps.supplier, mps.info, mps.filename from material AS m

    INNER JOIN material_price AS mp ON m.material_id = mp.material_id
    INNER JOIN material_price_summary AS mps ON mp.material_price_summary_id = mps.id

    WHERE material_brand IN :material_brand AND material_mode IN :material_mode LIMIT 1000;
    
    """)
    result = engine.execute(sql,{"material_brand":(brandModeInfo.get("brand")),"material_mode":(brandModeInfo.get("mode"))}).all()
    xheaders = u'规格型号,物料名称,类型,品牌,价格,报价数量,供应商,备注,原始文件'.split(',')
    xls_lines = [xheaders]
    logger.info(len(result))
    for item in result:
        logger.info(item)
        xls_lines.append(
            [item[0], item[1], item[2], item[3], item[4], item[5], item[6],
             item[7], item[8]])

    wb = openpyxl.Workbook()
    ws = wb.active
    for line in xls_lines:
        ws.append(line)
    wb.save("/opt/excel/out_brand_price.xlsx")
    send_email("supplier", "/opt/excel/out_brand_price.xlsx")

#保存报价信息
def saveMaterialPriceList(list, material_price_summary_id):
    for mPrice in list:
        r1 = session.query(Material).filter(Material.material_number == mPrice["number"]).all()
        if len(r1) > 0:
            print(r1[0].material_number)
            mp = MaterialPrice(price=mPrice["price"],
                               material_price_summary_id = material_price_summary_id,
                               material_id=r1[0].material_id,
                               amount = mPrice["amount"],)
            session.add(mp)
        else:
            print("FUCK it is none")
    session.commit()


#处理星奇系统里面的报价明细
def handleQuoteFile(filename, email):
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=172.17.0.239,1433;'
                          'Database=AIS20230524185151;'
                          'Uid=sa;'
                          'PWD=xqerp!@#2023;'
                          'Trusted_Connection=no;'
                          'TDS_Version=8.0')
    cursor = conn.cursor()

    sql_detail = '''
    
    SELECT TOP 100 material_l.*,  material.FNUMBER, materialpurchase.FMINPRICE,  materialpurchase.FMAXPRICE, supplier_l.FNAME AS sFNAME
      FROM dbo.T_BD_MATERIAL_L AS material_l 
    INNER JOIN dbo.T_BD_MATERIALPURCHASE AS materialpurchase ON materialpurchase.FMATERIALID = material_l.FMATERIALID
    INNER JOIN dbo.T_BD_MATERIAL AS material ON material.FMATERIALID = material_l.FMATERIALID

    INNER JOIN dbo.T_PUR_POORDERENTRY as poorderentry ON poorderentry.FMATERIALID = material_l.FMATERIALID
    INNER JOIN dbo.T_PUR_POORDER AS pur_poorder ON pur_poorder.FID = poorderentry.FID
    INNER JOIN dbo.T_BD_SUPPLIER AS supplier ON supplier.FSUPPLIERID = pur_poorder.FSUPPLIERID
    INNER JOIN dbo.T_BD_SUPPLIER_L AS supplier_l ON supplier_l.FSUPPLIERID = supplier.FSUPPLIERID

    WHERE material_l.FSPECIFICATION = ? ;
    
    '''

    print("开始处理excel文件"+filename)
    xheaders = u'FPKID,FMATARIALID,物料名称,规格型号,料号,最高价格,最低价格,供应商,购买次数'.split(',')
    xls_lines = [xheaders]


    file_name = filename
    refer_excel = openpyxl.load_workbook(file_name)
    #获取第一个sheet表格
    sheet1 = refer_excel['sheet']
    material_list = []
    for row in range(2, sheet1.max_row + 1):
        material_list.append({
            "name": (sheet1.cell(row=row, column=1)).value,
            "number": (sheet1.cell(row=row, column=2)).value,
        })
    #print(material_list)
    for material in material_list:
        #print(material["name"])
        number = material["number"]
        #print(number)
        rec = cursor.execute(sql_detail, number)
        datalist = rec.fetchall()
        if len(datalist) > 0:
            line = datalist[0]
            print(line)
            xls_lines.append(
                [line.FPKID, line.FMATERIALID, line.FNAME, number, line.FNUMBER, line.FMAXPRICE, line.FMINPRICE,
                 line.sFNAME, len(datalist)])
        else:
            xls_lines.append(["","",material["name"],number,"","","","",""])

    wb = openpyxl.Workbook()
    ws = wb.active
    for line in xls_lines:
        ws.append(line)
    cursor.close()
    conn.close()
    wb.save("/opt/excel/out_price.xlsx")

    email_name = "supplier"
    if email:
        email_name = email.replace('@xingqikeji.com','')
    send_email(email_name, "/opt/excel/out_price.xlsx")




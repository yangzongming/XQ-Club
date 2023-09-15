# -*- coding: utf-8 -*-
import re
import sys
import datetime
import requests
import pymssql
import time, os, tarfile
import openpyxl
import faulthandler
import pyodbc
from bottle import route, run ,template
from email.mime.application import MIMEApplication
import smtplib
from email.mime.multipart import MIMEMultipart
from email.header import Header

@route('/<mail>/<fid>')
def index(mail,fid):
    print(fid)
    craw_requestion_detail(fid)
    send_email(mail,fid)
    return template('<b>你好，你的请购询价单已发到你的邮箱里了 {{fid}}</b>!', fid=fid)

def text_db():
    coon = pymssql.connect(host='172.17.0.239',
                           user='sa',
                           password='xqerp!@#2023',
                           database='PYDB',
                           charset='utf-8')
    if coon:
        print("数据库连接成功了")
    else:
        print("faild")

def text_odbc():

    #AIS20230524185151
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=172.17.0.239,1433;'
                          'Database=AIS20230524185151;'
                          'Uid=sa;'
                          'PWD=xqerp!@#2023;'
                          'Trusted_Connection=no;'
                          'TDS_Version=8.0')
    if conn:
        print("sucessfull")
        cursor = conn.cursor()
        recSet = cursor.execute("select TOP 100 * from dbo.T_PUR_POORDER poorder inner join dbo.T_BD_SUPPLIER_L supplier on poorder.FSUPPLIERID = supplier.FSUPPLIERID "
                                "").fetchall()
        for one in recSet:
            print(u"订单FID=" + str(one.FID))
            print(u"订单号="+str(one.FBILLNO))
            print(u"供应商ID="+str(one.FSUPPLIERID))
            print(u"采购日期="+str(one.FDATE))
            print(u"简称："+str(one.FSHORTNAME))
            print("=========================")
        cursor.close()
        conn.close()

#获取采购订单明细
def craw_poorder_detail(fid):
    tmp_fid = fid
    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=172.17.0.239,1433;'
                          'Database=AIS20230524185151;'
                          'Uid=sa;'
                          'PWD=xqerp!@#2023;'
                          'Trusted_Connection=no;'
                          'TDS_Version=8.0')
    cursor = conn.cursor()
    recSet = cursor.execute("SELECT TOP 50 * from dbo.T_PUR_POORDERENTRY po_order_entry WHERE FID = " + str(fid)).fetchall()
    for one in recSet:
        print(u"物料编码:"+str(one.FMATERIALID))
        print(u"采购数量:" + str(one.FQTY))
        print(u"备注:"+str(one.FNOTE))

#获取采购申请表明细
def craw_requestion_detail(fid):
    #操作Excel
    file_name = "/opt/excel/price.xlsx"
    refer_excel = openpyxl.load_workbook(file_name)
    sheet1 = refer_excel['比价表']

    # 定义一个请购单数据dic
    pr_dic = {}
    pr_dic['no'] = fid  # 请购单号
    pr_dic['request_date'] = '2023-08-26'  # 请购日期
    pr_dic['request_person'] = 'leo yang'  # 请购人员
    pr_dic['pr_project'] = 'H2-天一阁'
    pr_dic['items'] = []

    conn = pyodbc.connect('Driver={SQL Server};'
                          'Server=172.17.0.239,1433;'
                          'Database=AIS20230524185151;'
                          'Uid=sa;'
                          'PWD=xqerp!@#2023;'
                          'Trusted_Connection=no;'
                          'TDS_Version=8.0')
    cursor = conn.cursor()
    sql = ''' SELECT * FROM dbo.T_PUR_REQUISITION AS pur_requesition WHERE pur_requesition.FBILLNO = ?  '''
    recSet = cursor.execute(sql,fid)

    # 处理料号
    limit = 6
    row_number = 0
    for one in recSet:
        print(str(one.FID))
        print(str(one.FBILLNO))
        print(str(one.FAPPROVEDATE))
        FBILLNO = str(one.FBILLNO)
        sql_detail = ''' SELECT pur_requestion.FID,   
                         material.FMATERIALID,material.F_XQZD_TEXT2,material.F_XQZD_TZBB, material_l.FSPECIFICATION, 
                         material_l.FNAME,material.FNUMBER, reqentry.FAPPROVEQTY ,
                         materialpurchase.FMINPRICE,materialpurchase.FMAXPRICE 
                         FROM dbo.T_PUR_REQUISITION AS pur_requestion 
                         INNER JOIN dbo.T_PUR_REQENTRY AS reqentry on pur_requestion.FID = reqentry.FID
                         INNER JOIN dbo.T_BD_MATERIAL AS material ON reqentry.FMATERIALID = material.FMATERIALID
                         INNER JOIN dbo.T_BD_MATERIAL_L AS material_l ON material_l.FMATERIALID = material.FMATERIALID
                         INNER JOIN dbo.T_BD_MATERIALPURCHASE AS materialpurchase ON materialpurchase.FMATERIALID = material.FMATERIALID
                         WHERE pur_requestion.FBILLNO = ?
                         ORDER BY pur_requestion.FCREATEDATE; '''
        rec = cursor.execute(sql_detail, FBILLNO)
        list = rec.fetchall()
        sheet1.insert_rows(7, len(list))
        for line in list:
            sheet1['A' + str(limit + row_number)] = row_number + 1
            sheet1['B' + str(limit + row_number)] = pr_dic['pr_project']
            sheet1['C' + str(limit + row_number)] = line.FNUMBER
            sheet1['D' + str(limit + row_number)] = line.FNAME
            sheet1['F' + str(limit + row_number)] = line.FSPECIFICATION
            sheet1['G' + str(limit + row_number)] = line.FAPPROVEQTY
            sheet1['H' + str(limit + row_number)] = line.FMAXPRICE
            sheet1['I' + str(limit + row_number)] = line.FMINPRICE
            row_number += 1
    #写数据
    sheet1['A3'] = u'采购申请单编号：' + pr_dic['no']
    sheet1['R3'] = u'期望交期' + pr_dic['request_date']
    refer_excel.save( "/opt/excel/" + fid + '.xlsx')


def text_excel(source_file):
    print("excel start")
    file_name = "/opt/craws/" + source_file + ".xlsx"
    refer_excel = openpyxl.load_workbook(file_name)
    sheet1 = refer_excel['导出表格']
    brand_model_dict = {}
    for row in range(2, sheet1.max_row + 1):
        brand = (sheet1.cell(row = row,column=1)).value
        model = (sheet1.cell(row = row,column=2)).value
        brand_model_dict[model] = brand
    print('brand_model_dict= ' + str(brand_model_dict))

def send_email(mail,fid):
    mail_host = "smtp.126.com"
    mail_user = "lao_yang_cool@126.com"
    mail_pass = "KVLXBNGMJSQSJBPT"

    sender = 'lao_yang_cool@126.com'
    receivers = [ mail + '@xingqikeji.com']

    message = MIMEMultipart()
    message['From'] = Header(u'请购单','utf-8')
    message['To'] =  Header(u'测试','utf-8')

    att1 = MIMEApplication(open('/opt/excel/' + str(fid) + '.xlsx', 'rb').read())
    att1["Content-Type"] = 'application/octet-stream'
    att1.add_header('Content-Disposition', 'attachment',
                    filename=(str(fid) + '.xlsx'))
    message.attach(att1)

    subject = u"请购询价单----" + str(fid)
    message['Subject'] = Header(subject, 'utf-8')
    try:
        smtpObj = smtplib.SMTP()
        smtpObj.connect(mail_host, 25)
        smtpObj.login(mail_user, mail_pass)
        smtpObj.sendmail(sender, receivers, message.as_string())
        print("mail send sucessfull")
    except smtplib.SMTPException:
        print("mail send error")

def craw_swagelok_info():
    print('begin swagelok')
    search_url = 'https://products.swagelok.com.cn/zh/%e6%89%80%e6%9c%89%e4%ba%a7%e5%93%81/%e6%8e%a5%e5%a4%b4/c/100?q=%3Arelevance%3Acategory%3A154&viewType=list&searchMode=&minisiteLanguage=&minisiteIssrlId='
    tmp_html = requests.get(search_url, headers=headers2())
    print(tmp_html)


def headers2():
    headers={}
    headers['User-Agent']='PostmanRuntime/7.26.8'
    headers['Accept']='*/*'
    headers['Accept-Encoding']='gzip, deflate, br'
    headers['Accept-Language']='zh-CN,zh;q=0.9'
    headers['Cache-Control']='no-cache'
    headers['Connection'] = 'keep-alive'
    headers['Host']='www.qixintong.cn'
    headers['Cookie']='Hm_lvt_8ade5034b029fd3234640e22eb8ac190=1642506651; hy_data_2020_id=17e6d0684646cf-036d0e8a0b6af9-133f6452-1296000-17e6d068465779; hy_data_2020_js_sdk=%7B%22distinct_id%22%3A%2217e6d0684646cf-036d0e8a0b6af9-133f6452-1296000-17e6d068465779%22%2C%22site_id%22%3A1896%2C%22user_company%22%3A1943%2C%22props%22%3A%7B%7D%2C%22device_id%22%3A%2217e6d0684646cf-036d0e8a0b6af9-133f6452-1296000-17e6d068465779%22%7D; _ga=GA1.2.1905398445.1642506651; _gid=GA1.2.130876695.1642506651; utype=1; sessionid=9swfbin9mlw96uyk9qdbzhf54d2ugbg0; tel=13522938301; _gat_gtag_UA_157572919_1=1; Hm_lpvt_8ade5034b029fd3234640e22eb8ac190=1642595683'
    return headers

if __name__ == '__main__':
    def application(environ, start_response):
        start_response('200 OK', [('Content-Type', 'text/html')])
        return ['<h1>Hello world!</h1>']
    run(host='localhost', port=8082)